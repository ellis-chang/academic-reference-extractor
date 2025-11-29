"""
Excel Output Module
Creates structured Excel files with author information
"""

import os
from typing import List, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


class ExcelOutputGenerator:
    """Generate Excel output files with extracted reference data"""
    
    def __init__(self):
        self.workbook = None
        self.main_sheet = None
        
        # Style definitions
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        self.data_font = Font(size=10)
        self.data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alternating row colors
        self.alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
    def create_workbook(self) -> Workbook:
        """Create a new workbook"""
        self.workbook = Workbook()
        self.main_sheet = self.workbook.active
        self.main_sheet.title = "Author Information"
        return self.workbook
    
    def write_reference_data(self, data: List[Dict[str, Any]], output_path: str):
        """
        Write reference data to Excel file
        
        Args:
            data: List of dictionaries with reference and author data
            output_path: Path for output Excel file
        """
        self.create_workbook()
        ws = self.main_sheet
        
        # Define columns
        columns = [
            ('Paper Title', 50),
            ('Year', 8),
            ('Chapter', 12),
            ('First Author Name', 25),
            ('First Author Affiliation', 35),
            ('First Author Department', 25),
            ('First Author Email', 30),
            ('Last Author Name', 25),
            ('Last Author Affiliation', 35),
            ('Last Author Department', 25),
            ('Last Author Email', 30),
            ('Citation Key', 15),
            ('Data Confidence', 12)
        ]
        
        # Write headers
        for col_idx, (header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Write data rows
        for row_idx, entry in enumerate(data, 2):
            row_data = [
                entry.get('paper_title', ''),
                entry.get('year', ''),
                entry.get('chapter', ''),
                entry.get('first_author_name', ''),
                entry.get('first_author_affiliation', ''),
                entry.get('first_author_department', ''),
                entry.get('first_author_email', ''),
                entry.get('last_author_name', ''),
                entry.get('last_author_affiliation', ''),
                entry.get('last_author_department', ''),
                entry.get('last_author_email', ''),
                entry.get('citation_key', ''),
                entry.get('confidence', '')
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.data_font
                cell.alignment = self.data_alignment
                cell.border = self.thin_border
                
                # Alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = self.alt_fill
        
        # Add summary sheet
        self._add_summary_sheet(data)
        
        # Save workbook
        self.workbook.save(output_path)
        return output_path
    
    def _add_summary_sheet(self, data: List[Dict[str, Any]]):
        """Add a summary sheet with statistics"""
        ws = self.workbook.create_sheet("Summary")
        
        # Calculate statistics
        total_refs = len(data)
        refs_with_first_author = sum(1 for d in data if d.get('first_author_affiliation'))
        refs_with_last_author = sum(1 for d in data if d.get('last_author_affiliation'))
        refs_with_emails = sum(1 for d in data if d.get('first_author_email') or d.get('last_author_email'))
        
        # Count by chapter
        chapters = {}
        for d in data:
            ch = d.get('chapter', 'Unknown')
            chapters[ch] = chapters.get(ch, 0) + 1
        
        # Write summary
        summary_data = [
            ('Extraction Summary', ''),
            ('', ''),
            ('Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('', ''),
            ('Total References', total_refs),
            ('References with First Author Affiliation', refs_with_first_author),
            ('References with Last Author Affiliation', refs_with_last_author),
            ('References with Email Addresses', refs_with_emails),
            ('', ''),
            ('Coverage Rate (First Author)', f'{refs_with_first_author/total_refs*100:.1f}%' if total_refs > 0 else 'N/A'),
            ('Coverage Rate (Last Author)', f'{refs_with_last_author/total_refs*100:.1f}%' if total_refs > 0 else 'N/A'),
            ('', ''),
            ('References by Chapter', ''),
        ]
        
        for chapter, count in sorted(chapters.items()):
            summary_data.append((chapter, count))
        
        # Write data
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True if row_idx == 1 else False)
            ws.cell(row=row_idx, column=2, value=value)
        
        # Format
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
    
    def export_to_csv(self, data: List[Dict[str, Any]], output_path: str):
        """Export data to CSV format"""
        df = pd.DataFrame(data)
        
        # Rename columns for clarity
        column_mapping = {
            'paper_title': 'Paper Title',
            'year': 'Year',
            'chapter': 'Chapter',
            'first_author_name': 'First Author Name',
            'first_author_affiliation': 'First Author Affiliation',
            'first_author_department': 'First Author Department',
            'first_author_email': 'First Author Email',
            'last_author_name': 'Last Author Name',
            'last_author_affiliation': 'Last Author Affiliation',
            'last_author_department': 'Last Author Department',
            'last_author_email': 'Last Author Email',
            'citation_key': 'Citation Key',
            'confidence': 'Data Confidence'
        }
        
        df = df.rename(columns=column_mapping)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        return output_path


def create_output_data(references, first_author_info, last_author_info) -> List[Dict[str, Any]]:
    """
    Create output data structure from references and author info
    
    Args:
        references: List of Reference objects
        first_author_info: Dict mapping reference index to AuthorInfo
        last_author_info: Dict mapping reference index to AuthorInfo
        
    Returns:
        List of dictionaries ready for Excel output
    """
    output_data = []
    
    for i, ref in enumerate(references):
        first_info = first_author_info.get(i)
        last_info = last_author_info.get(i)
        
        # Calculate average confidence
        confidences = []
        if first_info and first_info.confidence > 0:
            confidences.append(first_info.confidence)
        if last_info and last_info.confidence > 0:
            confidences.append(last_info.confidence)
        avg_confidence = sum(confidences) / len(confidences) if confidences else 0
        
        entry = {
            'paper_title': ref.title,
            'year': ref.year,
            'chapter': ref.chapter,
            'citation_key': ref.citation_key,
            'first_author_name': first_info.name if first_info else ref.first_author,
            'first_author_affiliation': first_info.affiliation if first_info else '',
            'first_author_department': first_info.department if first_info else '',
            'first_author_email': first_info.email if first_info else '',
            'last_author_name': last_info.name if last_info else ref.last_author,
            'last_author_affiliation': last_info.affiliation if last_info else '',
            'last_author_department': last_info.department if last_info else '',
            'last_author_email': last_info.email if last_info else '',
            'confidence': f'{avg_confidence:.0%}' if avg_confidence > 0 else ''
        }
        
        output_data.append(entry)
    
    return output_data


if __name__ == "__main__":
    # Test with sample data
    test_data = [
        {
            'paper_title': 'Deep Learning',
            'year': '2016',
            'chapter': 'Chapter 10',
            'first_author_name': 'Ian Goodfellow',
            'first_author_affiliation': 'Google Brain',
            'first_author_department': '',
            'first_author_email': '',
            'last_author_name': 'Yoshua Bengio',
            'last_author_affiliation': 'Université de Montréal',
            'last_author_department': 'MILA',
            'last_author_email': 'yoshua.bengio@umontreal.ca',
            'citation_key': "Goodfellow '16",
            'confidence': '85%'
        },
        {
            'paper_title': 'Attention Is All You Need',
            'year': '2017',
            'chapter': 'Chapter 12',
            'first_author_name': 'Ashish Vaswani',
            'first_author_affiliation': 'Google Research',
            'first_author_department': '',
            'first_author_email': '',
            'last_author_name': 'Illia Polosukhin',
            'last_author_affiliation': '',
            'last_author_department': '',
            'last_author_email': '',
            'citation_key': "Vaswani '17",
            'confidence': '70%'
        }
    ]
    
    generator = ExcelOutputGenerator()
    output_path = '/tmp/test_output.xlsx'
    generator.write_reference_data(test_data, output_path)
    print(f"Test file created: {output_path}")
